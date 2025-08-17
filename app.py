# app.py — DHF Streamlit UI (Preview + ZIP Download) — full file
# -----------------------------------------------------------------------------

import os
import io
import json
import zipfile
import typing as t
from datetime import datetime

import pandas as pd
import streamlit as st
import requests

# ---------------- Constants & Secrets ----------------
TBD = "TBD - Human / SME input"

BACKEND_URL = st.secrets.get("BACKEND_URL", "http://localhost:8080")
BACKEND_TOKEN = st.secrets.get("BACKEND_TOKEN", "dev-token")

# Row caps (preview + Excel export)
HA_MAX_ROWS = int(os.getenv("HA_MAX_ROWS", "50"))
DVP_MAX_ROWS = int(os.getenv("DVP_MAX_ROWS", "50"))
TM_MAX_ROWS  = int(os.getenv("TM_MAX_ROWS", "50"))

# NEW: cap how many requirements we send to backend
REQ_MAX = int(os.getenv("REQ_MAX", "50"))

# Optional Google Drive upload (still supported)
DEFAULT_DRIVE_FOLDER_ID = st.secrets.get("DRIVE_FOLDER_ID", "")
OUTPUT_DIR = st.secrets.get("OUTPUT_DIR", os.path.abspath("./streamlit_outputs"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------- Optional Google Drive (pydrive2) ----------------
_HAS_DRIVE = True
try:
    from pydrive2.auth import GoogleAuth
    from pydrive2.drive import GoogleDrive
except Exception:
    _HAS_DRIVE = False


def init_drive() -> t.Optional["GoogleDrive"]:
    if not _HAS_DRIVE:
        return None
    svc_json = st.secrets.get("SERVICE_ACCOUNT_JSON")
    if not svc_json:
        return None
    os.makedirs(".secrets", exist_ok=True)
    svc_path = os.path.join(".secrets", "service_account.json")
    with open(svc_path, "w", encoding="utf-8") as f:
        f.write(svc_json)
    gauth = GoogleAuth()
    try:
        gauth.LoadServiceAccountCredentials(svc_path)
    except Exception:
        gauth.settings.update({
            'client_config_backend': 'service',
            'service_config': {
                'client_json_file_path': svc_path,
                'client_user_email': json.loads(svc_json).get('client_email', ''),
            },
            'oauth_scope': ['https://www.googleapis.com/auth/drive']
        })
        gauth.ServiceAuth()
    return GoogleDrive(gauth)


def drive_upload_bytes(drive: "GoogleDrive", folder_id: str, filename: str, data: bytes) -> str:
    file = drive.CreateFile({"title": filename, "parents": [{"id": folder_id}]})
    file.content = io.BytesIO(data)
    file.Upload()
    return file["id"]

# ---------------- Helpers ----------------
def normalize_requirements(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for c in df.columns:
        lc = str(c).strip().lower()
        if lc in {"requirement id", "req id", "requirement_id"}:
            rename_map[c] = "Requirement ID"
        elif lc in {"verification id", "verification_id", "verif id"}:
            rename_map[c] = "Verification ID"
        elif lc in {"requirement", "requirements", "requirement text", "requirement_desc", "requirement description"}:
            rename_map[c] = "Requirements"
    df = df.rename(columns=rename_map)
    for col in ["Requirement ID", "Verification ID", "Requirements"]:
        if col not in df.columns:
            df[col] = None
    return df[["Requirement ID", "Verification ID", "Requirements"]].copy()


def call_backend(endpoint: str, payload: dict) -> dict:
    url = f"{BACKEND_URL.rstrip('/')}{endpoint}"
    headers = {"Authorization": f"Bearer {BACKEND_TOKEN}", "Content-Type": "application/json"}
    r = requests.post(url, headers=headers, json=payload, timeout=1200)
    if r.status_code >= 400:
        raise RuntimeError(f"Backend error {r.status_code}: {r.text[:500]}")
    return r.json()


def head_cap(df: pd.DataFrame, n: int) -> pd.DataFrame:
    return df.head(n).copy() if isinstance(df, pd.DataFrame) and not df.empty else df


def _excel_style_sheet(ws, header_fill, wrap=True, col_widths=None, row_height=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    align = Alignment(horizontal="left", vertical="center", wrapText=wrap)
    # Header row (1)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = align
    # Rows
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        if row_height:
            ws.row_dimensions[r].height = row_height
        for cell in ws[r]:
            cell.alignment = align
    # Column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    # Freeze panes: 2nd row & 4th column
    ws.freeze_panes = ws.cell(row=2, column=4)

def df_to_excel_bytes(df: pd.DataFrame, kind: str) -> bytes:
    """
    kind: "ha" | "dvp" | "tm" -> apply per-sheet formatting.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    green_fill = PatternFill(start_color="A8D08D", end_color="A8D08D", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write data
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row.values))

    # Column-letter helper
    def col_letter(idx: int) -> str:
        from openpyxl.utils import get_column_letter
        return get_column_letter(idx)

    # Default widths
    col_widths = {col_letter(i+1): 15 for i in range(len(df.columns))}
    row_height = None

    if kind == "ha":
        # Special: hazardous_situation & sequence_of_events -> 40; risk_control -> 100
        name_to_width = {
            "hazardous_situation": 40,
            "sequence_of_events": 40,
            "risk_control": 100,
        }
        for i, c in enumerate(df.columns):
            w = name_to_width.get(c, None)
            if w:
                col_widths[col_letter(i+1)] = w

    elif kind == "dvp":
        # Requirements & test_procedure 60; row height 60
        name_to_width = {
            "requirements": 60,
            "test_procedure": 60,
        }
        for i, c in enumerate(df.columns):
            if c in name_to_width:
                col_widths[col_letter(i+1)] = name_to_width[c]
        row_height = 60

    elif kind == "tm":
        # Requirements & HA Risk Control = 100
        name_to_width = {
            "Requirements": 100,
            "HA Risk Control": 100,
        }
        for i, c in enumerate(df.columns):
            w = name_to_width.get(c, None)
            if w:
                col_widths[col_letter(i+1)] = w

    _excel_style_sheet(ws, header_fill=green_fill, wrap=True, col_widths=col_widths, row_height=row_height)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def basic_guardrails_df(df: pd.DataFrame, required_cols: t.List[str]) -> pd.DataFrame:
    issues = []
    for i, row in df.iterrows():
        for c in required_cols:
            val = str(row.get(c, "")).strip()
            if not val or val in {"NA", TBD}:
                issues.append((i, c, "Missing or TBD"))
    return pd.DataFrame(issues, columns=["row_index", "column", "issue"]) if issues else pd.DataFrame(columns=["row_index", "column", "issue"])


# ---------------- Metric helpers ----------------
TBD_CANON = {TBD, "TBD", "NA", "N/A", ""}

def _pct(num: int, den: int) -> float:
    if den <= 0:
        return 0.0
    return round(100.0 * num / den, 1)

def _not_tbd_val(x) -> bool:
    s = str(x or "").strip()
    return bool(s) and s not in TBD_CANON

def _norm_text_series(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
         .str.lower()
         .str.replace(r"[^a-z0-9 ]+", " ", regex=True)
         .str.replace(r"\s+", " ", regex=True)
         .str.strip()
    )

def ha_metrics(ha_df: pd.DataFrame) -> dict:
    if ha_df is None or ha_df.empty:
        return {"field_complete": 0.0, "diversity": 0.0}

    keys = [
        "risk_to_health", "hazard", "hazardous_situation", "harm",
        "sequence_of_events", "severity_of_harm", "p0", "p1", "poh",
        "risk_index", "risk_control",
    ]
    filled = 0
    for k in keys:
        if k in ha_df.columns:
            filled += int(ha_df[k].apply(_not_tbd_val).sum())
    field_complete = _pct(filled, len(ha_df) * len(keys))

    cols = ["risk_to_health", "hazard", "hazardous_situation", "sequence_of_events", "harm"]
    sub = ha_df[cols].copy()
    for c in cols:
        if c in sub.columns:
            sub[c] = _norm_text_series(sub[c])
    uniq = sub.drop_duplicates().shape[0]
    diversity = _pct(uniq, len(ha_df))

    return {"field_complete": field_complete, "diversity": diversity}

def dvp_metrics(dvp_df: pd.DataFrame) -> dict:
    if dvp_df is None or dvp_df.empty:
        return {"field_complete": 0.0, "sev_sample": 0.0}

    keys = ["verification_method", "sample_size", "test_procedure", "acceptance_criteria"]
    filled = 0
    for k in keys:
        if k in dvp_df.columns:
            filled += int(dvp_df[k].apply(_not_tbd_val).sum())
    field_complete = _pct(filled, len(dvp_df) * len(keys))

    def _ok_sample(x: str) -> bool:
        try:
            return int(str(x).strip()) > 0
        except Exception:
            return False
    sev_sample = _pct(int(dvp_df["sample_size"].apply(_ok_sample).sum()), len(dvp_df))

    return {"field_complete": field_complete, "sev_sample": sev_sample}

def tm_metrics(tm_df: pd.DataFrame) -> dict:
    if tm_df is None or tm_df.empty:
        return {"map_pct": 0.0, "field_complete": 0.0}

    keys = ["Requirement ID", "Requirements", "Risk ID", "Risk to Health",
            "HA Risk Control", "Verification ID", "Verification Method", "Acceptance Criteria"]
    filled = 0
    for k in keys:
        if k in tm_df.columns:
            filled += int(tm_df[k].apply(_not_tbd_val).sum())
    field_complete = _pct(filled, len(tm_df) * len(keys))

    req_col = "Requirements"
    rc_col  = "HA Risk Control"
    ac_col  = "Acceptance Criteria"

    SYN = {
        "flow": {"rate", "accuracy", "volume"},
        "occlusion": {"block", "blockage"},
        "alarm": {"alert", "beep"},
        "leakage": {"leak"},
        "air": {"bubble", "embolism"},
        "embolism": {"air", "bubble"},
        "vibration": {"vibrate"},
        "temperature": {"thermal", "heat"},
        "pressure": {"kpa", "bar"},
    }

    def _tokens(s: str) -> set:
        toks = [
            t for t in _norm_text_series(pd.Series([s])).iloc[0].split()
            if len(t) >= 3 and t not in {"shall", "system", "device", "patient", "the"}
        ]
        return set(toks)

    def _expand(tokens: set) -> set:
        out = set(tokens)
        for t in list(tokens):
            for k, vals in SYN.items():
                if t == k or t in vals:
                    out.add(k)
                    out.update(vals)
        return out

    def _jaccard(a: set, b: set) -> float:
        if not a or not b:
            return 0.0
        inter = a & b
        union = a | b
        return len(inter) / len(union)

    mapped = 0
    for _, r in tm_df.iterrows():
        req = _expand(_tokens(str(r.get(req_col, ""))))
        rc  = _expand(_tokens(str(r.get(rc_col, ""))))
        ac  = _expand(_tokens(str(r.get(ac_col, ""))))
        ok = False
        if req and ((req & rc) or (req & ac)):
            ok = True
        else:
            if max(_jaccard(req, rc), _jaccard(req, ac)) >= 0.08:
                ok = True
        mapped += int(ok)

    map_pct = _pct(mapped, len(tm_df))
    return {"map_pct": map_pct, "field_complete": field_complete}

def _metric_line(label: str, value: float, threshold: float) -> None:
    color = "#198754" if value >= threshold else "#C1121F"
    st.markdown(
        f"""
        <div style="font-size:22px;margin-top:8px;">
          <span style="opacity:0.75">{label} (Threshold: {int(threshold)}%)</span><br/>
          <span style="font-size:44px;font-weight:800;color:{color};">{value:.1f}%</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_metrics(ha_df: pd.DataFrame, dvp_df: pd.DataFrame, tm_df: pd.DataFrame):
    ha = ha_metrics(ha_df)
    dvp = dvp_metrics(dvp_df)
    tm  = tm_metrics(tm_df)

    st.subheader("Evaluation Metrics")
    st.markdown(
        "Objective: these metrics quickly summarize **coverage**, **consistency**, and "
        "**requirement↔control mapping** quality across the generated documents."
    )

    # ---- First row (requested order)
    r1c1, r1c2, r1c3 = st.columns(3)
    with r1c1:
        _metric_line("HA • Field Completeness", ha["field_complete"], 80)
    with r1c2:
        _metric_line("DVP • Field Completeness", dvp["field_complete"], 80)
    with r1c3:
        _metric_line("TM • Field Completeness", tm["field_complete"], 80)

    # ---- Second row (requested order)
    r2c1, r2c2, r2c3 = st.columns(3)
    with r2c1:
        _metric_line("HA • Scenario Diversity", ha["diversity"], 70)
    with r2c2:
        _metric_line("DVP • Severity↔Sample Alignment", dvp["sev_sample"], 70)
    with r2c3:
        _metric_line("TM • Req↔RiskControl Mapping", tm["map_pct"], 80)

    st.markdown(
        '<div style="margin-top:12px;font-weight:800;color:#0b63d1;">'
        'Note: Please involve Medical Device SMEs for final review of these documents before approval.'
        '</div>',
        unsafe_allow_html=True,
    )

# ---------------- UI ----------------
st.set_page_config(page_title="DHF Automation – Infusion Pump", layout="wide")

# Header with puzzle icon + two images to the right
col_logo, col_title, col_img1, col_img2 = st.columns([0.4, 2.6, 1.4, 1.4])
with col_logo:
    # Place your icon at: streamlit_assets/puzzle.png
    try:
        st.image("streamlit_assets/puzzle.png", width=56)
    except Exception:
        st.write("")
with col_title:
    st.markdown(
        '<div style="font-size:50px; font-weight:800; color:#6E42C1; line-height:1.05;">'
        'DHF Automation – Infusion Pump'
        '</div>',
        unsafe_allow_html=True,
    )
with col_img1:
    try:
        st.image("streamlit_assets/Infusion1.jpg", use_container_width=True)
    except Exception:
        pass
with col_img2:
    try:
        st.image("streamlit_assets/Infusion.jpg", use_container_width=True)
    except Exception:
        pass

st.caption("Requirements → Hazard Analysis → DVP → TM")

# Inputs
st.markdown("**Provide Product Requirements** (choose one):")
colA, colB = st.columns(2)
with colA:
    uploaded = st.file_uploader("Upload Product Requirements (Excel .xlsx)", type=["xlsx", "xls"])
with colB:
    sample = "Requirement ID,Verification ID,Requirements\nPR-001,VER-001,System shall ..."
    pasted = st.text_area("Paste as CSV (with headers)", value="", height=180, placeholder=sample)

# Top, visible Generate button (green)
run_btn = st.button("🟩 Generate DHF Packages", type="primary")

if run_btn:
    # -------- Parse Requirements --------
    with st.spinner("Parsing requirements..."):
        if uploaded is not None:
            try:
                req_df = pd.read_excel(uploaded)
            except Exception as e:
                st.error(f"Failed to read Excel: {e}")
                st.stop()
        elif pasted.strip():
            try:
                req_df = pd.read_csv(io.StringIO(pasted))
            except Exception as e:
                st.error(f"Failed to parse pasted CSV: {e}")
                st.stop()
        else:
            st.error("Please upload an Excel file or paste CSV text.")
            st.stop()

        req_df = normalize_requirements(req_df)

    total_reqs = len(req_df)
    st.success(f"Loaded {total_reqs} requirements.")
    st.dataframe(head_cap(req_df, 20), use_container_width=True)

    # -------- Limit how many requirements are SENT to backend --------
    req_df_limited = req_df.head(REQ_MAX).copy()
    if total_reqs > REQ_MAX:
        st.info(f"Sending only the first {REQ_MAX} requirements to backend (set REQ_MAX env var to change).")

    # -------- Call Backend: HA --------
    with st.spinner("Running Hazard Analysis (backend)..."):
        ha_payload = {
            "requirements": [
                {
                    "Requirement ID": r["Requirement ID"],
                    "Verification ID": r.get("Verification ID"),
                    "Requirements": r.get("Requirements"),
                }
                for _, r in req_df_limited.iterrows()
            ]
        }
        ha_resp = call_backend("/hazard-analysis", ha_payload)
        ha_rows = ha_resp.get("ha", [])
        ha_df = pd.DataFrame(ha_rows)

    # Fill TBDs
    def fill_tbd(df: pd.DataFrame, cols: t.List[str]) -> pd.DataFrame:
        out = df.copy()
        for c in cols:
            if c not in out.columns:
                out[c] = TBD
            out[c] = out[c].fillna(TBD)
            out.loc[out[c].astype(str).str.strip().eq(""), c] = TBD
            out.loc[out[c].astype(str).str.upper().eq("NA"), c] = TBD
        return out

    ha_df = fill_tbd(
        ha_df,
        [
            "requirement_id", "risk_id", "risk_to_health", "hazard", "hazardous_situation",
            "harm", "sequence_of_events", "severity_of_harm", "p0", "p1", "poh", "risk_index", "risk_control",
        ],
    )
    st.subheader(f"Hazard Analysis (preview, first {HA_MAX_ROWS})")
    st.dataframe(head_cap(ha_df, HA_MAX_ROWS), use_container_width=True)

    # -------- Call Backend: DVP --------
    with st.spinner("Generating Design Verification Protocol (backend)..."):
        dvp_payload = {
            "requirements": ha_payload["requirements"],  # already limited
            "ha": ha_rows,
        }
    try:
        dvp_resp = call_backend("/dvp", dvp_payload)
        dvp_rows = dvp_resp.get("dvp", [])
    except Exception as e:
        st.error(f"DVP backend error: {e}")
        st.stop()
    dvp_df = pd.DataFrame(dvp_rows)
    dvp_df = fill_tbd(dvp_df, ["verification_id", "requirement_id", "requirements",
                               "verification_method", "sample_size", "test_procedure", "acceptance_criteria"])
    st.subheader(f"Design Verification Protocol (preview, first {DVP_MAX_ROWS})")
    st.dataframe(head_cap(dvp_df, DVP_MAX_ROWS), use_container_width=True)

    # -------- Call Backend: TM --------
    with st.spinner("Building Trace Matrix (backend)..."):
        tm_payload = {
            "requirements": ha_payload["requirements"],
            "ha": ha_rows,
            "dvp": dvp_rows,
        }
        tm_resp = call_backend("/tm", tm_payload)  # backend returns {"ok":true,"tm":[...]}
        tm_rows = tm_resp.get("tm", [])
        tm_df = pd.DataFrame(tm_rows)

    TM_ORDER = [
        "Requirement ID",
        "Requirements",
        "Requirement (Yes/No)",
        "Risk ID",
        "Risk to Health",
        "HA Risk Control",
        "Verification ID",
        "Verification Method",
        "Acceptance Criteria",
    ]
    tm_df = tm_df.reindex(columns=TM_ORDER)
    tm_df = fill_tbd(tm_df, TM_ORDER)

    st.subheader(f"Trace Matrix (preview, first {TM_MAX_ROWS})")
    st.dataframe(head_cap(tm_df, TM_MAX_ROWS), use_container_width=True)

    # -------- Evaluation Metrics (ordered as requested) --------
    render_metrics(ha_df, dvp_df, tm_df)

    # -------- Prepare styled Excel exports (capped for preview) --------
    ha_export_cols = [
        "requirement_id", "risk_id", "risk_to_health", "hazard", "hazardous_situation",
        "harm", "sequence_of_events", "severity_of_harm", "p0", "p1", "poh", "risk_index", "risk_control",
    ]
    dvp_export_cols = ["verification_id", "requirement_id", "requirements", "verification_method",
                       "sample_size", "test_procedure", "acceptance_criteria"]
    tm_export_cols = TM_ORDER[:]  # same as preview

    ha_bytes  = df_to_excel_bytes(head_cap(ha_df[ha_export_cols], HA_MAX_ROWS), kind="ha")
    dvp_bytes = df_to_excel_bytes(head_cap(dvp_df[dvp_export_cols], DVP_MAX_ROWS), kind="dvp")
    tm_bytes  = df_to_excel_bytes(head_cap(tm_df[tm_export_cols], TM_MAX_ROWS), kind="tm")

    # Save locally
    with open(os.path.join(OUTPUT_DIR, "Hazard_Analysis.xlsx"), "wb") as f:
        f.write(ha_bytes)
    with open(os.path.join(OUTPUT_DIR, "Design_Verification_Protocol.xlsx"), "wb") as f:
        f.write(dvp_bytes)
    with open(os.path.join(OUTPUT_DIR, "Trace_Matrix.xlsx"), "wb") as f:
        f.write(tm_bytes)

    # -------- One-click ZIP Download --------
    st.subheader("Download DHF Package")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Hazard_Analysis.xlsx", ha_bytes)
        zf.writestr("Design_Verification_Protocol.xlsx", dvp_bytes)
        zf.writestr("Trace_Matrix.xlsx", tm_bytes)
    zip_buf.seek(0)

    clicked = st.download_button(
        "⬇️ Download DHF package (3 Excel files, ZIP)",
        data=zip_buf.getvalue(),
        file_name=f"DHF_Package_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
        mime="application/zip",
        type="primary"
    )
    if clicked:
        st.markdown(
            '<div style="margin-top:8px;font-weight:800;">'
            'DHF documents downloaded successfully - Hazard Analysis, Design Verification Protocol, Trace Matrix'
            '</div>',
            unsafe_allow_html=True,
        )

    # -------- Optional Google Drive upload --------
    if DEFAULT_DRIVE_FOLDER_ID:
        with st.spinner("Uploading files to Google Drive..."):
            drive = init_drive()
            if drive:
                try:
                    ha_id = drive_upload_bytes(drive, DEFAULT_DRIVE_FOLDER_ID, "Hazard_Analysis.xlsx", ha_bytes)
                    dvp_id = drive_upload_bytes(drive, DEFAULT_DRIVE_FOLDER_ID, "Design_Verification_Protocol.xlsx", dvp_bytes)
                    tm_id = drive_upload_bytes(drive, DEFAULT_DRIVE_FOLDER_ID, "Trace_Matrix.xlsx", tm_bytes)
                    st.info("Uploaded to Google Drive (file IDs shown below).")
                    st.json({
                        "Hazard_Analysis.xlsx": ha_id,
                        "Design_Verification_Protocol.xlsx": dvp_id,
                        "Trace_Matrix.xlsx": tm_id,
                    })
                except Exception as e:
                    st.warning(f"Drive upload failed: {e}")
            else:
                st.info("Drive not initialized (missing SERVICE_ACCOUNT_JSON secret).")
